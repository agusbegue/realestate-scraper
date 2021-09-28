import io
from xlsxwriter.workbook import Workbook
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from main.utils import constants as c
# from scrapy_app.utils.business import LEN_PROPERTIES


class FileCreator:

    def __init__(self, props):
        self.props = props

    def get_data(self):
        output = io.BytesIO()
        workbook = Workbook(output, {'in_memory': True})

        self._create_summary(self.props, workbook)
        for prop in self.props:
            self._write_sheet(prop, workbook)

        workbook.close()
        output.seek(0)
        data = output.read()
        output.close()

        return data

    def _create_summary(self, props, workbook):

        sheet = workbook.add_worksheet('Summary')

        header_format = workbook.add_format({'bold': 1,
                                             'align': 'center',
                                             'valign': 'vcenter'})

        sheet.merge_range(0, 0, 1, 0, 'Address', header_format)

        for j in range(3):
            '''LEN_PROPERTIES'''
            sheet.merge_range(0, 5*j + 1, 0, 5*j + 5, 'Testigo ' + str(j+1), header_format)
            sheet.write(1, 5*j+1, 'Link', header_format)
            sheet.write(1, 5*j+2, 'Dist.', header_format)
            sheet.write(1, 5*j+3, 'Price', header_format)
            sheet.write(1, 5*j+4, 'Area', header_format)
            sheet.write(1, 5*j+5, '€/m²', header_format)

        for i, prop in enumerate(props):
            sheet.write(2+i, 0, prop.address)

            posts = {post.index: post for post in prop.post_set.all()}
            for j, post in posts.items():
                sheet.write(2+i, 5*j-4, c.BASE_LINK + post.link)
                sheet.write(2+i, 5*j-3, post.distance)
                sheet.write(2+i, 5*j-2, post.price)
                sheet.write(2+i, 5*j-1, post.area)
                sheet.write(2+i, 5*j, round(post.price/post.area, 2))

    def _write_sheet(self, prop, workbook):

        sheet = workbook.add_worksheet('Property ' + str(len(workbook.worksheets())))

        # map_boolean = lambda f: "Si" if f else "No"
        header_format = workbook.add_format({'bold': 1})
        sheet.write(0, 0, prop.address + ' - ' + prop.type, header_format)
        sheet.write(1, 0, 'Link', header_format)
        sheet.write(1, 1, 'Distance', header_format)
        sheet.write(1, 2, 'Price', header_format)
        sheet.write(1, 3, 'Area', header_format)
        sheet.write(1, 4, '€/m²', header_format)
        sheet.write(1, 5, 'Address', header_format)

        for i, post in enumerate(prop.post_set.all()):
            sheet.write(2+i, 0, c.BASE_LINK + post.link)
            sheet.write(2+i, 1, post.distance)
            sheet.write(2+i, 2, post.price)
            sheet.write(2+i, 3, post.area)
            sheet.write(2+i, 4, round(post.price/post.area, 2))
            sheet.write(2+i, 5, post.address)


class FileReader:

    def __init__(self, job):
        self.job = job

    def read(self, file):
        sheet = self._validate(file)
        if sheet:
            return self._create_objects(sheet)
        return

    def _validate(self, file):

        if file.size > c.MAX_FILE_SIZE:
            self.job.status = c.FAILED
            self.job.details = 'Max file size exceded {} expected less than {}'.format(c.filesize(file.size),
                                                                                       c.filesize(c.MAX_FILE_SIZE))
            return

        try:
            wb = load_workbook(file)
        except InvalidFileException:
            self.job.status = c.FAILED
            self.job.details = 'Invalid file, error while opening'
            return

        if len(wb.sheetnames) > 1:
            self.job.status = c.FAILED
            self.job.details = 'Input with more than one sheet'
            return

        sheet = wb[wb.sheetnames[0]]
        return sheet

    def _map_boolean(self, value):
        if isinstance(value, str):
            try:
                int_val = int(value)
                return int_val != 0
            except ValueError:
                return value.lower().strip() == 'yes'
        return value != 0

    def _parse_float(self, coord):
        # ver que onda cuando este vacio
        if isinstance(coord, str):
            coord.replace(',', '.')
            return float(coord)
        return coord

    def _create_objects(self, sheet):

        invalid = []
        valid = []
        for i in range(2, sheet.max_row + 1):
            info = {}
            try:
                info['type'] = sheet.cell(row=i, column=c.TYPE).value
                info['address'] = " ".join([sheet.cell(row=i, column=c.ADDRESS).value,
                                            sheet.cell(row=i, column=c.CITY).value,
                                            sheet.cell(row=i, column=c.PROVINCE).value])
                info['area'] = self._parse_float(sheet.cell(row=i, column=c.AREA).value)
                info['parking'] = self._map_boolean(sheet.cell(row=i, column=c.PARKING).value)
                info['lift'] = self._map_boolean(sheet.cell(row=i, column=c.LIFT).value)
                info['latitude'] = self._parse_float(sheet.cell(row=i, column=c.LATITUDE).value)
                info['longitude'] = self._parse_float(sheet.cell(row=i, column=c.LONGITUDE).value)
                valid.append(info)
            except ValueError:
                info['valid'] = False
                invalid.append(info)
            except Exception:
                info['valid'] = False
                invalid.append(info)

        if len(valid) == 0:
            self.job.status = c.FAILED
            self.job.details = 'No valid properties on file'
        elif len(invalid) > 0:
            self.job.details = '{} invalid properties in file'.format(len(invalid))
        self.job.records = len(valid) + len(invalid)

        return valid + invalid


