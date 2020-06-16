import json
from django.utils import timezone
from django.db import models
from django.core.validators import ValidationError
from django.contrib.auth.models import AbstractUser

from scrapyd_api import ScrapydAPI
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from main.settings import SCRAPYD_URL, MAX_FILE_SIZE
from main import constants as c


class User(AbstractUser):

    email = models.EmailField('Email address', unique=True, null=False)
    REQUIRED_FIELDS = ['email']


class ScrapyJob(models.Model):

    id = models.AutoField(primary_key=True)
    status = models.CharField(max_length=10, default=c.PENDING)
    task = models.CharField(max_length=25, default=c.NOT_DEFINED)
    details = models.CharField(max_length=50, null=True)
    date = models.DateTimeField(default=timezone.now)
    records = models.IntegerField(null=True)
    http_status = models.IntegerField(null=True)
    user = models.ForeignKey(User, on_delete=models.CASCADE)

    def start(self, file):
        properties = self._validate_input(file)
        self._send()
        self.save()
        if properties:
            self._save_properties(properties)


    def _send(self):
        scrapyd = ScrapydAPI(SCRAPYD_URL)
        # self.task = scrapyd.schedule('default', 'idealista', job_key=self.id)
        self.task = '012345'

    def _validate_input(self, file):

        if file.size > MAX_FILE_SIZE:
            self.status = c.FAILED
            self.details = 'Max file size exceded {} expected less than {}'.format(c.filesize(file.size),
                                                                                   c.filesize(MAX_FILE_SIZE))
            return

        try:
            wb = load_workbook(file)
        except InvalidFileException:
            self.status = c.FAILED
            self.details = 'Invalid file, error while opening'
            return

        if len(wb.sheetnames) > 1:
            self.status = c.FAILED
            self.details = 'Input with more than one sheet'
            return

        sheet = wb[wb.sheetnames[0]]

        map_boolean = lambda val: val.lower().strip() == 'si'
        invalidas = []
        validas = []
        for i in range(2, sheet.max_row+1):
            info = {}
            try:
                info['tipo'] = sheet.cell(row=i, column=c.TIPO).value
                info['direccion'] = " ".join([sheet.cell(row=i, column=c.DIRECCION).value,
                                              sheet.cell(row=i, column=c.MUNICIPIO).value,
                                              sheet.cell(row=i, column=c.PROVINCIA).value])
                info['superficie'] = sheet.cell(row=i, column=c.SUPERFICIE).value
                info['parking'] = map_boolean(sheet.cell(row=i, column=c.PARKING).value)
                info['ascensor'] = map_boolean(sheet.cell(row=i, column=c.ASCENSOR).value)
                propiedad = Property(**info)
                validas.append(propiedad)
            except ValidationError:
                propiedad = Property(valid=False, **info)
                invalidas.append(propiedad)
            except Exception:
                #fila en blanco
                pass

        if len(validas) == 0:
            self.status = c.FAILED
            self.details = 'No valid properties on file'
        elif len(invalidas) > 0:
            self.details = 'Some invalid properties in file'
        self.records = len(validas) + len(invalidas)

        return validas + invalidas

    def _save_properties(self, properties):
        for property in properties:
            property.job_id = self.id
        Property.objects.bulk_create(properties)

    @property
    def row_color(self):
        if self.status == 'Pending' or self.status == 'Running':
            return 'LightYellow'
        elif self.status == 'Failed':
            return 'LightCoral'
        else:
            return 'PaleGreen'


class Property(models.Model):

    type = models.CharField(max_length=30)
    address = models.CharField(max_length=100)
    surface = models.FloatField()
    parking = models.BooleanField(default=False)
    elevator = models.BooleanField(default=False)
    valid = models.BooleanField(default=True)
    job = models.ForeignKey(ScrapyJob, on_delete=models.CASCADE)

    def clean(self):
        if False:
            raise ValidationError('Incorrect field format')
        elif False:
            raise SyntaxError('Blank fields')

    def to_file(self):
        import io

        from django.http.response import HttpResponse

        from xlsxwriter.workbook import Workbook

        def your_view(request):
            output = io.BytesIO()

            workbook = Workbook(output, {'in_memory': True})
            worksheet = workbook.add_worksheet()
            worksheet.write(0, 0, 'Hello, world!')
            workbook.close()

            output.seek(0)

            response = HttpResponse(output.read(),
                                    content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            response['Content-Disposition'] = "attachment; filename=test.xlsx"

            output.close()

            return response

    @property
    def row_color(self):
        if self.valid:
            return 'PaleGreen'
        else:
            return 'LightCoral'

class Statistics(models.Model):

    property = models.ForeignKey(Property, on_delete=models.CASCADE)
    price = models.FloatField()


class UserData(models.Model):
    pass

from django.core.validators import FileExtensionValidator
class FileModel(models.Model):
    file = models.FileField('Excel file', validators=[FileExtensionValidator(allowed_extensions=['xlsx', 'csv'])])

