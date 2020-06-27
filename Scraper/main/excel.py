from django.core.validators import ValidationError

import io
from xlsxwriter.workbook import Workbook
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from main.settings import MAX_FILE_SIZE
from main import constants as c


class FileCreator:

    def __init__(self, props, bulk):
        self.props = props
        self.bulk = bulk

    def get_data(self):
        output = io.BytesIO()
        workbook = Workbook(output, {'in_memory': True})

        if self.bulk:
            for prop in self.props:
                self._write_sheet(prop, workbook)
        else:
            self._write_sheet(self.props, workbook)

        workbook.close()
        output.seek(0)
        data = output.read()
        output.close()

        return data

    def _write_sheet(self, prop, workbook):

        sheet = workbook.add_worksheet(prop.address)

        map_boolean = lambda f: "Si" if f else "No"
        sheet.write(0, 0, "Type")
        sheet.write(0, 1, prop.type)
        sheet.write(1, 0, "Address")
        sheet.write(1, 1, prop.address)
        sheet.write(2, 0, "Surface")
        sheet.write(2, 1, prop.surface)
        sheet.write(3, 0, "Parking")
        sheet.write(3, 1, map_boolean(prop.parking))
        sheet.write(4, 0, "Elevator")
        sheet.write(4, 1, map_boolean(prop.elevator))


class FileReader:

    def __init__(self, job):
        self.job = job

    def read(self, file):
        sheet = self._validate(file)
        if sheet:
            return self._create_objects(sheet)
        return

    def _validate(self, file):

        if file.size > MAX_FILE_SIZE:
            self.job.status = c.FAILED
            self.job.details = 'Max file size exceded {} expected less than {}'.format(c.filesize(file.size),
                                                                                       c.filesize(MAX_FILE_SIZE))
            return

        try:
            wb = load_workbook(file)
        except InvalidFileException:
            self.job.status = c.FAILED
            self.job.details = 'Invalid file, error while opening'
            return

        if len(wb.sheetnames) > 1:
            self.job.status = c.FAILED
            self.job.details = 'Input with more than one sheet'
            return

        sheet = wb[wb.sheetnames[0]]
        return sheet

    def _create_objects(self, sheet):

        map_boolean = lambda val: val.lower().strip() == 'si'
        invalid = []
        valid = []
        for i in range(2, sheet.max_row + 1):
            info = {}
            try:
                info['tipo'] = sheet.cell(row=i, column=c.TIPO).value
                info['direccion'] = " ".join([sheet.cell(row=i, column=c.DIRECCION).value,
                                              sheet.cell(row=i, column=c.MUNICIPIO).value,
                                              sheet.cell(row=i, column=c.PROVINCIA).value])
                info['superficie'] = sheet.cell(row=i, column=c.SUPERFICIE).value
                info['parking'] = map_boolean(sheet.cell(row=i, column=c.PARKING).value)
                info['ascensor'] = map_boolean(sheet.cell(row=i, column=c.ASCENSOR).value)
                valid.append(info)
            except ValidationError:
                info['valid'] = False
                invalid.append(info)
            except Exception:
                # fila en blanco
                pass

        if len(valid) == 0:
            self.job.status = c.FAILED
            self.job.details = 'No valid properties on file'
        elif len(invalid) > 0:
            self.job.details = '{} invalid properties in file'.format(len(invalid))
        self.job.records = len(valid) + len(invalid)

        return valid + invalid
